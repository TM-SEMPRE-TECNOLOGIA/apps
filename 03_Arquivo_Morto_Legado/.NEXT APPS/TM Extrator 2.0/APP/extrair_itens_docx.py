# Requisitos:
#   pip install python-docx openpyxl
#
# Uso:
#   python extrair_itens_docx.py
#
# O que faz:
# - Tkinter: seleciona .docx e escolhe onde salvar .xlsx
# - Lê SOMENTE tabelas cujo cabeçalho (1ª linha) contenha "Itens"
# - Extrai (Codigo, Quantidade), sem unidade
# - Gera um log (.txt) no mesmo diretório do .xlsx

import os
import re
from collections import defaultdict
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


CODE_RE = re.compile(r"^\s*\d+(?:\.\d+)?\s*$")  # aceita 17.4, 13.12, 2.24, 19.37 etc.


def iter_block_items(parent):
    """
    Itera sobre parágrafos e tabelas na ordem em que aparecem.
    """
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    
    # Se o parent tem 'element', acessamos o corpo. 
    # Document (o objeto doc) tem .element.body.
    if hasattr(parent, "element") and hasattr(parent.element, "body"):
        parent_elm = parent.element.body
    elif hasattr(parent, "_element"):
        parent_elm = parent._element
    else:
        return

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)



def norm(s: str) -> str:
    return (s or "").replace("\xa0", " ").strip()


def is_itens_table(table) -> bool:
    # Tabela-alvo: 
    # Formato antigo: primeira linha contém "Itens" em qualquer célula
    # Novo formato: cabeçalho contém "(ITEM XX.XX)"
    # Memorial: cabeçalho contém "Itens", mas as linhas seguintes são resumo [Código, Desc, Qtd, Un]
    if not table.rows:
        return False, "none", None
    first_row_texts = [norm(c.text) for c in table.rows[0].cells]
    header_text = " ".join(first_row_texts)
    
    # Check new format
    m_new = re.search(r"\(\s*I*TEM\s+([\d.]+)\s*\)", header_text, re.IGNORECASE)
    if m_new:
        return True, "new", m_new.group(1).strip()
        
    # Check if is a potential 'Itens' table
    is_itens_header = any(t.lower() == "itens" for t in first_row_texts)
    if is_itens_header:
        # Se tiver mais de uma linha, e a segunda linha tiver o padrão de código e quantidade,
        # vamos tratar como 'old' para extrair todas as linhas de dados.
        if len(table.rows) > 1:
            second_row = [norm(c.text) for c in table.rows[1].cells]
            if len(second_row) >= 3:
                code = second_row[0]
                qty = second_row[2]
                if CODE_RE.match(code) and re.match(r"^[\d.,]+$", qty):
                    # Se for o formato de resumo de 4 colunas (Memorial), 
                    # tratamos como 'old' para ler todas as linhas da tabela.
                    return True, "old", None

        return True, "old", None
        
    return False, "none", None



def pick_quantity_from_row(cells_text):
    """
    Preferência:
    - coluna 3 (índice 2), pois geralmente é a quantidade
    Fallback:
    - procurar o primeiro número no texto da linha (1,00 / 277,50 / 10 / 2,5 etc.)
    """
    if len(cells_text) >= 3:
        q = norm(cells_text[2])
        if q and q.upper() != "#N/D":
            return q

    joined = " ".join(cells_text)
    m = re.search(r"(\d{1,3}(?:\.\d{3})*,\d+|\d+,\d+|\d+)", joined)
    return m.group(1) if m else ""


def extract_code_qty(docx_path: str):
    doc = Document(docx_path)

    results = []
    logs = []
    memorials = {}
    item_info = {}   # code -> {desc, unit, col_headers, measurements[]}
    itens_tables = 0
    
    in_calculation_section = False
    target_header = "MEMORIAL DE CÁLCULO E ITENS:"

    for block in iter_block_items(doc):
        # Detectar mudança de seção
        if isinstance(block, Paragraph):
            txt = norm(block.text).upper()
            if target_header in txt:
                in_calculation_section = True
            continue
            
        if not isinstance(block, Table):
            continue
            
        table = block
        is_itens, fmt, code_detected = is_itens_table(table)
        if not is_itens:
            continue

        # SE estivermos na seção de Memorial, ignorar tabelas tipo 'new' (são resumos do que já lemos)
        # Manter leitura do fmt 'memorial' (o resumo de 4 colunas) para meta-informação
        if in_calculation_section and fmt == "new":
            continue

        itens_tables += 1

        if fmt == "old":
            # Formato antigo: pula cabeçalho e analisa cada linha
            for r_i, row in enumerate(table.rows[1:], start=2):
                cells = [norm(c.text) for c in row.cells]
                if not cells:
                    logs.append((itens_tables, r_i, "skip_empty_row", ""))
                    continue

                code = norm(cells[0])
                if not code or code.upper() == "#N/D":
                    logs.append((itens_tables, r_i, "skip_code_empty_or_ND", ""))
                    continue

                if not CODE_RE.match(code):
                    logs.append((itens_tables, r_i, "skip_code_invalid", code))
                    continue

                qty = pick_quantity_from_row(cells)
                if not qty or qty.upper() == "#N/D":
                    logs.append((itens_tables, r_i, "skip_qty_empty_or_ND", code))
                    continue

                results.append((code.strip(), qty.strip()))

        elif fmt == "new":
            # --- Capturar informações completas do item ---
            header_cell = norm(table.rows[0].cells[0].text)
            desc = re.sub(r"\s*\(I*TEM\s+[\d.]+\)\s*$", "", header_cell,
                          flags=re.IGNORECASE).strip()

            # Cabeçalhos de coluna (R1)
            col_headers = []
            unit = "?"
            if len(table.rows) > 1:
                col_headers = [norm(c.text) for c in table.rows[1].cells]
                for ct in col_headers:
                    um = re.search(r"TOTAL\s*\(([^)]+)\)", ct, re.IGNORECASE)
                    if um:
                        unit = um.group(1).strip()
                        break

            # Registrar info do item (primeiro encontro define colunas)
            if code_detected not in item_info:
                item_info[code_detected] = {
                    "desc": desc,
                    "unit": unit,
                    "col_headers": col_headers,
                    "measurements": []
                }

            # Separar linhas de dados, subtotal e total
            data_rows = []
            subtotal_row = None
            total_row = None

            for r_i, row in enumerate(table.rows[2:], start=3):
                cells = [norm(c.text) for c in row.cells]
                if not cells:
                    logs.append((itens_tables, r_i, "skip_empty_row_NEW", ""))
                    continue

                row_label = cells[0].lower()
                if "subtotal" in row_label:
                    subtotal_row = cells
                    continue
                if "total" in row_label:
                    total_row = cells
                    continue

                # Linha de dados
                qty = norm(cells[-1])
                if not qty or qty.upper() == "#N/D":
                    logs.append((itens_tables, r_i, "skip_qty_empty_or_ND_NEW", code_detected))
                    continue

                if re.match(r"^[\d.,]+$", qty):
                    results.append((code_detected, qty))
                    data_rows.append(cells)
                else:
                    logs.append((itens_tables, r_i, "skip_qty_invalid_NEW", qty))

            # Guardar medição completa para o memorial de cálculo
            item_info[code_detected]["measurements"].append({
                "table_idx": itens_tables,
                "data_rows": data_rows,
                "subtotal_row": subtotal_row,
                "total_row": total_row,
            })

        elif fmt == "memorial":
            # Quadro resumo do final do relatório
            cells = [norm(c.text) for c in table.rows[1].cells]
            code = cells[0]
            desc = cells[1]
            qty_doc = cells[2]
            unit = cells[3]
            memorials[code] = {
                "desc": desc,
                "qty_doc": qty_doc,
                "unit": unit
            }

    meta = {
        "tables_total": itens_tables, # Usamos o contador de tabelas úteis
        "itens_tables": itens_tables,
        "rows_extracted": len(results),
        "rows_ignored": len(logs),
        "ignored_details": logs,
        "memorials": memorials,
        "item_info": item_info,
    }
    return results, meta


def save_xlsx(rows, meta, out_path: str):
    wb = Workbook()

    def fmt_pt(val):
        """Formata float para PT-BR: 1234.56 -> '1.234,56'"""
        return f"{val:,.2f}".replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")

    def natural_key(k):
        return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', k)]

    memorials_doc = meta.get("memorials", {})
    item_info = meta.get("item_info", {})

    # --- Aba 1: Itens (Todos) ---
    ws = wb.active
    ws.title = "Itens"
    ws.append(["Codigo", "Quantidade"])
    for code, qty in rows:
        ws.append([code, qty])
    ws.freeze_panes = "A2"
    _auto_width(ws)

    # --- Aba 2: Consolidado ---
    ws_cons = wb.create_sheet(title="Consolidado")
    consolidated_list = consolidate_rows(rows)
    ws_cons.append(["Codigo", "Quantidade Total"])
    for code, qty in consolidated_list:
        ws_cons.append([code, qty])
    ws_cons.freeze_panes = "A2"
    _auto_width(ws_cons)

    # --- Aba 3: MEMORIAL FINAL ---
    ws_final = wb.create_sheet(title="Memorial Final")
    has_doc_memorial = bool(memorials_doc)

    if has_doc_memorial:
        ws_final.append(["Código", "Descrição", "Qtd Script",
                         "Qtd Documento", "Unidade", "Status"])
    else:
        ws_final.append(["Código", "Descrição", "Quantidade Total", "Unidade"])

    # Calcular somas usando Total das tabelas (que já consideram descontos/multiplicadores)
    real_sums = {}
    for code, info in item_info.items():
        real_total = 0.0
        for meas in info["measurements"]:
            if meas["total_row"]:
                real_total += parse_pt_br_float(meas["total_row"][-1])
            elif meas["data_rows"]:
                for dr in meas["data_rows"]:
                    real_total += parse_pt_br_float(dr[-1])
        real_sums[code] = real_total

    # Fallback: somas brutas para itens sem item_info (formato antigo)
    raw_sums = defaultdict(float)
    for c, q in rows:
        raw_sums[c] += parse_pt_br_float(q)

    all_codes = set(raw_sums.keys()) | set(item_info.keys()) | set(memorials_doc.keys())
    sorted_final = sorted(all_codes, key=natural_key)

    for code in sorted_final:
        total_script = real_sums.get(code, raw_sums.get(code, 0.0))
        info = item_info.get(code, {})
        doc_mem = memorials_doc.get(code)

        desc = info.get("desc", doc_mem["desc"] if doc_mem else "")
        unit = info.get("unit", doc_mem["unit"] if doc_mem else "")

        if has_doc_memorial:
            qty_doc_str = doc_mem["qty_doc"] if doc_mem else "N/A"
            qty_doc_val = parse_pt_br_float(qty_doc_str) if doc_mem else 0.0
            diff = abs(total_script - qty_doc_val)
            status = "CONFERE" if diff < 0.01 else "DIVERGENTE"
            if not doc_mem:
                status = "SEM MEMORIAL NO DOC"
            ws_final.append([code, desc, fmt_pt(total_script),
                             qty_doc_str, unit, status])
        else:
            ws_final.append([code, desc, fmt_pt(total_script), unit])

    ws_final.freeze_panes = "A2"
    _auto_width(ws_final)

    # --- Abas 4+: Memorial de Cálculo por Item ---
    sorted_items = sorted(item_info.keys(), key=natural_key)

    for code in sorted_items:
        info = item_info[code]
        measurements = info["measurements"]
        col_headers = info.get("col_headers", [])
        desc = info.get("desc", "")
        unit = info.get("unit", "?")
        n_cols = len(col_headers) if col_headers else 2

        # Detectar se alguma tabela tem Total com modificador (X3, DIVIDIDO, etc.)
        has_special_total = False
        for meas in measurements:
            if meas["total_row"]:
                label = meas["total_row"][0].lower()
                if label != "total" and "total" in label:
                    has_special_total = True
                    break

        sheet_name = f"MemCalc_{code}"[:31]
        ws_mc = wb.create_sheet(title=sheet_name)

        # R0: Cabeçalho com descrição e código (estilo MODELO LIMPO)
        header_row = [""] * n_cols
        header_row[0] = f"{desc} (ITEM {code})"
        ws_mc.append(header_row)

        # R1: Cabeçalhos de coluna
        if col_headers:
            ws_mc.append(col_headers)
        else:
            ws_mc.append(["REFERÊNCIA", f"TOTAL ({unit})"])

        all_data_count = 0

        if has_special_total:
            # --- MODO EMPILHADO: cada tabela com seu subtotal/total ---
            grand_total = 0.0
            for meas in measurements:
                for dr in meas["data_rows"]:
                    ws_mc.append(dr)
                    all_data_count += 1

                # Subtotal da tabela individual (se existir)
                if meas["subtotal_row"]:
                    ws_mc.append(meas["subtotal_row"])

                # Total da tabela individual (com modificador)
                if meas["total_row"]:
                    ws_mc.append(meas["total_row"])
                    grand_total += parse_pt_br_float(meas["total_row"][-1])
                elif meas["data_rows"]:
                    for dr in meas["data_rows"]:
                        grand_total += parse_pt_br_float(dr[-1])

            # Total geral
            total_geral = [""] * n_cols
            total_geral[0] = "TOTAL GERAL"
            total_geral[-1] = fmt_pt(grand_total)
            ws_mc.append(total_geral)

        else:
            # --- MODO SIMPLES: todas as linhas juntas, subtotal + total ---
            grand_total = 0.0
            for meas in measurements:
                for dr in meas["data_rows"]:
                    ws_mc.append(dr)
                    all_data_count += 1
                    grand_total += parse_pt_br_float(dr[-1]) if dr else 0.0

            # Subtotal
            sub_row = [""] * n_cols
            sub_row[0] = "Subtotal"
            if n_cols >= 4:
                # Colunas de desconto: somar descontos
                discount_sum = 0.0
                for meas in measurements:
                    for dr in meas["data_rows"]:
                        if len(dr) >= 4:
                            discount_sum += parse_pt_br_float(dr[-2])
                sub_row[-2] = fmt_pt(discount_sum)
            sub_row[-1] = fmt_pt(grand_total)
            ws_mc.append(sub_row)

            # Total (usando os totals reais das tabelas, que consideram descontos)
            real_total = 0.0
            for meas in measurements:
                if meas["total_row"]:
                    real_total += parse_pt_br_float(meas["total_row"][-1])
                elif meas["data_rows"]:
                    for dr in meas["data_rows"]:
                        real_total += parse_pt_br_float(dr[-1])

            total_row = [""] * n_cols
            total_row[0] = "Total"
            total_row[-1] = fmt_pt(real_total)
            ws_mc.append(total_row)

        _auto_width(ws_mc)

    wb.save(out_path)


def _auto_width(ws):
    """Ajusta a largura das colunas baseada no conteúdo."""
    for col_cells in ws.columns:
        max_len = 0
        column = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_len + 2, 50)


def parse_pt_br_float(s: str) -> float:
    """Converte '1.234,56' ou '100' para float."""
    try:
        # Remove pontos de milhar e troca vírgula por ponto
        clean = s.replace(".", "").replace(",", ".")
        return float(clean)
    except ValueError:
        return 0.0


def consolidate_rows(rows):
    """
    Soma as quantidades agrupando pelo código.
    Retorna lista ordenada de (codigo, qtd_formatada).
    """
    sums = defaultdict(float)

    for code, qty_str in rows:
        val = parse_pt_br_float(qty_str)
        sums[code] += val

    # Ordenação natural para códigos como 17.1, 17.2, 17.10...
    def natural_key(k):
        return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', k)]

    sorted_codes = sorted(sums.keys(), key=natural_key)

    consolidated = []
    for code in sorted_codes:
        total = sums[code]
        # Formata float para PT-BR com separadores de milhar
        # Ex: 1234.56 -> "1.234,56"
        txt = f"{total:,.2f}".replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
        consolidated.append((code, txt))

    return consolidated


def save_log(meta, rows, docx_path, out_xlsx_path):
    log_path = os.path.splitext(out_xlsx_path)[0] + "_log.txt"

    # Gera dados consolidados
    consolidated = consolidate_rows(rows)

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("LOG - Extração de Tabelas 'Itens'\n")
        f.write(f"Arquivo: {os.path.basename(docx_path)}\n")
        f.write(f"Data/hora: {datetime.now().isoformat(timespec='seconds')}\n\n")

        f.write(f"Tabelas totais no DOCX: {meta['tables_total']}\n")
        f.write(f"Tabelas identificadas como 'Itens': {meta['itens_tables']}\n")
        f.write(f"Linhas extraídas (total): {meta['rows_extracted']}\n")
        f.write(f"Linhas ignoradas: {meta['rows_ignored']}\n\n")

        if meta["ignored_details"]:
            f.write("Detalhes ignorados (tabela, linha, motivo, valor):\n")
            for t_i, r_i, reason, val in meta["ignored_details"]:
                f.write(f"- T{t_i} L{r_i}: {reason} {val}\n")
            f.write("\n")

        f.write("=== LOG CONSOLIDADO (Soma por Código) ===\n")
        f.write("(Codigo | Quantidade Total)\n")
        for c, q in consolidated:
            f.write(f"{c} | {q}\n")
        f.write("\n")

        f.write("=== LOG DETALHADO (Extração Original) ===\n")
        f.write("(Codigo | Quantidade)\n")
        for c, q in rows:
            f.write(f"{c} | {q}\n")

    return log_path


def run_gui():
    root = tk.Tk()
    root.withdraw()

    docx_path = filedialog.askopenfilename(
        title="Selecione o arquivo Word (.docx)",
        filetypes=[("Word (.docx)", "*.docx")]
    )
    if not docx_path:
        return

    out_xlsx = filedialog.asksaveasfilename(
        title="Salvar saída em planilha (.xlsx)",
        defaultextension=".xlsx",
        filetypes=[("Excel (.xlsx)", "*.xlsx")]
    )
    if not out_xlsx:
        return

    try:
        rows, meta = extract_code_qty(docx_path)

        if not rows and not meta.get("memorials"):
            messagebox.showwarning(
                "Sem dados",
                "Nenhuma medição ou memorial foi encontrado no documento."
            )
            return

        save_xlsx(rows, meta, out_xlsx)
        log_path = save_log(meta, rows, docx_path, out_xlsx)

        messagebox.showinfo(
            "Concluído",
            f"Planilha criada:\n{out_xlsx}\n\nLog criado:\n{log_path}"
        )
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")


if __name__ == "__main__":
    run_gui()
