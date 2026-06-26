import sys
from openpyxl import load_workbook

def analyze_styles(path):
    try:
        wb = load_workbook(path, data_only=True)
        print(f"Planilha: {path}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"\nAba: {sheet}")
            for r in range(1, 15):
                for c in range(1, 10):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None and r > 1: continue
                    
                    fill_color = "None"
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                        fill_color = cell.fill.fgColor.rgb
                    elif cell.fill and cell.fill.start_color and cell.fill.start_color.type == 'rgb':
                        fill_color = cell.fill.start_color.rgb
                    
                    font_color = "000000"
                    if cell.font and cell.font.color and cell.font.color.type == 'rgb':
                        font_color = cell.font.color.rgb
                        
                    bold = cell.font.bold if cell.font else False
                    print(f"[{r},{c}] V: {str(cell.value)[:20]} | Fill: {fill_color} | Font: {font_color} | Bold: {bold}")
    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    analyze_styles(sys.argv[1])
