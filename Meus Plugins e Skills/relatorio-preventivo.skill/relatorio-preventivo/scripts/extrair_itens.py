"""
extrair_itens.py — Etapa 1: extrai lista de itens de um relatório preventivo.
Uso: python3 extrair_itens.py <input.md> <output.xlsx>
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from parse_report import (parse_calc_occurrences, parse_itens_tables,
                           consolidate_occurrences, find_memorial_start, fmt)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, name='Calibri', size=10, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
    return c


def data_cell(ws, row, col, value, align='left', bold=False, bg='FFFFFF'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Calibri', size=9, bold=bold)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin_border()
    return c


def build_sheet(ws, title, items_list, qty_col_header='Quantidade'):
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:D1')
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, name='Calibri', size=12, color='FFFFFF')
    t.fill = PatternFill('solid', start_color='1F4E79')
    t.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 20
    for ci, h in enumerate(['Código', 'Descrição', qty_col_header, 'Unidade'], 1):
        header_cell(ws, 2, ci, h)

    for ri, item in enumerate(items_list, 3):
        bg = 'EBF3FB' if ri % 2 == 0 else 'FFFFFF'
        data_cell(ws, ri, 1, item['code'], align='center', bg=bg)
        data_cell(ws, ri, 2, item['desc'], bg=bg)
        data_cell(ws, ri, 3, item['qty_str'], align='center', bg=bg)
        data_cell(ws, ri, 4, item['unit'], align='center', bg=bg)
        ws.row_dimensions[ri].height = 28

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12


def main(md_path, out_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    memorial_start = find_memorial_start(lines)
    body_lines = lines[:memorial_start] if memorial_start else lines

    occurrences = parse_calc_occurrences(body_lines)
    items_info, _ = parse_itens_tables(body_lines)
    consolidated = consolidate_occurrences(occurrences)

    # Lista completa (ordem de aparição, todas as ocorrências)
    all_items = []
    seen = {}
    for occ in occurrences:
        code = occ['code']
        info = items_info.get(code, {})
        all_items.append({
            'code': code,
            'desc': info.get('desc', occ['title'].split('(')[0].strip()),
            'qty_str': fmt(occ['total_value']),
            'unit': info.get('unit', occ['total_unit'])
        })

    # Lista consolidada (ordem de primeira aparição, somada)
    consol_list = []
    for code, d in consolidated.items():
        info = items_info.get(code, {})
        consol_list.append({
            'code': code,
            'desc': info.get('desc', d['title'].split('(')[0].strip()),
            'qty_str': fmt(d['total_sum']),
            'unit': info.get('unit', d['total_unit'])
        })

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Lista Completa'
    build_sheet(ws1, 'LISTA DE ITENS – ORDEM DO RELATÓRIO', all_items)

    ws2 = wb.create_sheet('Consolidado')
    build_sheet(ws2, 'LISTA CONSOLIDADA – QUANTIDADES SOMADAS',
                consol_list, qty_col_header='Quantidade Total')

    wb.save(out_path)
    print(f'✅ Excel salvo: {out_path}')
    print(f'   Lista completa: {len(all_items)} linhas | Consolidado: {len(consol_list)} itens únicos')


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python3 extrair_itens.py <input.md> <output.xlsx>')
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
